from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent

EXCEL_FILE_PATH = BASE_DIR / "spreadsheet.xlsx"

EXCEL_HEADERS = [
    "Row",
    "Column Header 1",
    "Column Header 2",
    "Column Header 3",
]


def create_excel_with_headers(file_path=EXCEL_FILE_PATH, headers_list=EXCEL_HEADERS):
    wb = Workbook()
    ws = wb.active
    ws.title = "TD_0_to_LS_export_result"
    ws.append(headers_list)
    wb.save(filename=file_path)
    return True


def add_row_to_excel(row_list, file_path=EXCEL_FILE_PATH, headers=EXCEL_HEADERS):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    ws.append(row_list)

    # Optional: Adjust column widths for better readability
    # This is a simple auto-width, might not be perfect for all data
    for col_idx, header in enumerate(headers, 1):
        max_length = 0
        does_column_has_long_texts = (
            "link" in header.lower()
            or header.lower() == "data with annot"
            or header.lower() == "users creation result"
            or header.lower() == "data with predict"
        )
        for cell in ws[get_column_letter(col_idx)]:
            if does_column_has_long_texts:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass  # Handle cases where cell.value might be None
        adjusted_width = (max_length + 2) * 1.1  # Add some padding
        if does_column_has_long_texts:
            adjusted_width = (len(header) + 2) * 1.1
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(filename=file_path)
    return True
